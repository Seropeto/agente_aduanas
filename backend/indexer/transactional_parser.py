"""
Parser de documentación transaccional (Sprint AGENT-002).

Extrae texto de archivos .pdf / .xlsx / .csv recibidos como bytes y los segmenta
en chunks listos para embeddings. Para documentos tabulares (xlsx/csv) segmenta
por filas/bloques (cada fila = un registro). Para PDF (típicamente la DIN o fichas
técnicas) segmenta por tamaño respetando límites de palabra.

Cada chunk se prefija con una cabecera de operación para que el contexto de la
operación viaje dentro del propio vector:
    "[Operación: 2025-DIN-5582 | Origen: India] <contenido>"
"""
import csv
import io
import logging
import re

logger = logging.getLogger(__name__)

ALLOWED_EXTS = {".pdf", ".xlsx", ".csv"}

# Chunking de prosa (PDF): ~220 palabras, solape 30 — adecuado para fichas/DIN.
CHUNK_WORDS = 220
CHUNK_OVERLAP = 30
# Filas por chunk para tabulares (xlsx/csv): agrupa varias filas por fragmento.
ROWS_PER_CHUNK = 20


def _ext(filename: str) -> str:
    i = filename.rfind(".")
    return filename[i:].lower() if i >= 0 else ""


# ── Extracción de texto por formato ───────────────────────────────────────────

def _extract_pdf(raw: bytes) -> str:
    import pdfplumber
    parts = []
    with pdfplumber.open(io.BytesIO(raw)) as pdf:
        for page in pdf.pages:
            t = page.extract_text() or ""
            if t.strip():
                parts.append(t.strip())
    return "\n\n".join(parts)


def _extract_xlsx(raw: bytes) -> list[str]:
    """Devuelve una lista de filas como texto 'col1: v1 | col2: v2 | ...'."""
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    rows_text: list[str] = []
    for ws in wb.worksheets:
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        headers = [str(h).strip() if h is not None else f"col{i}" for i, h in enumerate(header or [])]
        for row in rows:
            cells = []
            for i, val in enumerate(row):
                if val is None or str(val).strip() == "":
                    continue
                col = headers[i] if i < len(headers) else f"col{i}"
                cells.append(f"{col}: {val}")
            if cells:
                rows_text.append(" | ".join(cells))
    wb.close()
    return rows_text


def _extract_csv(raw: bytes) -> list[str]:
    text = None
    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            text = raw.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        return []
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        return []
    headers = [h.strip() for h in rows[0]]
    out = []
    for row in rows[1:]:
        cells = []
        for i, val in enumerate(row):
            if not str(val).strip():
                continue
            col = headers[i] if i < len(headers) else f"col{i}"
            cells.append(f"{col}: {val}")
        if cells:
            out.append(" | ".join(cells))
    return out


# ── Chunking ──────────────────────────────────────────────────────────────────

def _chunk_prose(text: str) -> list[str]:
    text = re.sub(r"[ \t]+", " ", text).strip()
    words = text.split()
    if len(words) <= CHUNK_WORDS:
        return [text] if text else []
    chunks, start = [], 0
    while start < len(words):
        end = min(start + CHUNK_WORDS, len(words))
        chunks.append(" ".join(words[start:end]))
        nxt = end - CHUNK_OVERLAP
        start = nxt if nxt > start else end
    return chunks


def _chunk_rows(rows: list[str]) -> list[str]:
    """Agrupa filas tabulares en bloques (cada bloque = ROWS_PER_CHUNK filas)."""
    return [
        "\n".join(rows[i:i + ROWS_PER_CHUNK])
        for i in range(0, len(rows), ROWS_PER_CHUNK)
        if rows[i:i + ROWS_PER_CHUNK]
    ]


def parse_and_chunk(raw: bytes, filename: str, operation_id: str, origin_country: str) -> list[str]:
    """
    Extrae texto del archivo y devuelve una lista de chunks con cabecera de operación.
    Lanza ValueError si el formato no es soportado.
    """
    ext = _ext(filename)
    if ext not in ALLOWED_EXTS:
        raise ValueError(f"Formato no soportado: {ext}")

    if ext == ".pdf":
        base_chunks = _chunk_prose(_extract_pdf(raw))
    elif ext == ".xlsx":
        base_chunks = _chunk_rows(_extract_xlsx(raw))
    else:  # .csv
        base_chunks = _chunk_rows(_extract_csv(raw))

    header = f"[Operación: {operation_id} | Origen: {origin_country}] "
    return [header + c for c in base_chunks if c.strip()]
